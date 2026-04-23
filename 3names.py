import streamlit as st
import pandas as pd
import re
from rapidfuzz import fuzz
from openpyxl.styles import PatternFill
from io import BytesIO

# --- الجزء الأول: منطق المطابقة الأصلي (دقة 100%) ---
def normalize_name(name):
    if pd.isnull(name): return ""
    name = str(name).strip()
    name = name.replace("ه","ة").replace("أ","ا").replace("إ","ا").replace("آ","ا")
    name = name.replace("ى","ي").replace("ئ","ي")
    name = re.sub(r'(عبد)([^\s])', r'\1 \2', name)
    return " ".join(name.split()).lower()

def get_first_three_words(name):
    if pd.isnull(name) or name == "": return ""
    words = str(name).split()
    return " ".join(words[:3]) if len(words) >= 3 else " ".join(words)

# --- واجهة المستخدم (Streamlit) ---
st.set_page_config(page_title="نظام المطابقة المطور", layout="wide")
st.title("🔄 نظام المطابقة الذكي وجلب البيانات المتعددة")

# 1. رفع الملفات
col_f1, col_f2 = st.columns(2)
with col_f1:
    file_source = st.file_uploader("رفع ملف المصدر (مثلاً: ملف بصرة)", type=['xlsx'])
with col_f2:
    file_target = st.file_uploader("رفع ملف الهدف (مثلاً: ملف العقاري)", type=['xlsx'])

if file_source and file_target:
    df_s = pd.read_excel(file_source)
    df_t = pd.read_excel(file_target)

    st.markdown("### ⚙️ إعدادات الربط الديناميكي")
    
    # توزيع الإعدادات في أعمدة
    set1, set2, set3 = st.columns(3)
    
    with set1:
        st.info("الأعمدة الأساسية (المصدر)")
        name_col_s = st.selectbox("عمود الاسم في ملف المصدر", df_s.columns)
        school_col_s = st.selectbox("عمود المدرسة/القسم في ملف المصدر", df_s.columns)

    with set2:
        st.info("الأعمدة الأساسية (الهدف)")
        name_col_t = st.selectbox("عمود الاسم في ملف الهدف", df_t.columns)
        school_col_t = st.selectbox("عمود المدرسة في ملف الهدف", df_t.columns)

    with set3:
        st.info("جلب البيانات")
        # هنا التطوير: اختيار أي عدد من الأعمدة لجلبها (IBAN، رقم الهاتف، الخ)
        cols_to_fetch = st.multiselect(
            "اختر الأعمدة المراد نقلها من المصدر إلى الهدف عند التطابق",
            [c for c in df_s.columns if c not in [name_col_s, school_col_s]],
            help="يمكنك اختيار الـ Iban وأي أعمدة أخرى تريدها"
        )

    if st.button("🚀 بدء عملية المطابقة"):
        with st.spinner("جاري التحليل والمطابقة..."):
            
            # تحضير البيانات (نفس المنطق الأصلي)
            df_s["_norm_name"]   = df_s[name_col_s].apply(normalize_name)
            df_s["_three_word"]  = df_s["_norm_name"].apply(get_first_three_words)
            df_s["_norm_school"] = df_s[school_col_s].apply(normalize_name)

            df_t["_norm_name"]   = df_t[name_col_t].apply(normalize_name)
            df_t["_three_word"]  = df_t["_norm_name"].apply(get_first_three_words)
            df_t["_norm_school"] = df_t[school_col_t].apply(normalize_name)

            final_results = []

            # دورة المطابقة (نفس منطق الكود الأصلي تماماً)
            for _, row_t in df_t.iterrows():
                t_three  = row_t["_three_word"]
                t_school = row_t["_norm_school"]
                
                # البحث في المصدر بناءً على أول 3 كلمات
                candidates = df_s[df_s["_three_word"] == t_three]

                # بناء السجل الأساسي (يحتوي على كافة بيانات ملف الهدف الأصلي)
                # استبعاد الأعمدة المؤقتة التي تبدأ بـ "_"
                record = {k: v for k, v in row_t.to_dict().items() if not str(k).startswith('_')}

                if candidates.empty:
                    record["الاسم المطابق"] = ""
                    record["نسبة تطابق المدرسة"] = ""
                    record["الملاحظة"] = "❌ لا يوجد اسم مطابق"
                    for col in cols_to_fetch: record[f"المجلوب_{col}"] = ""
                else:
                    best_score, best_row_s = 0, None
                    for _, row_s in candidates.iterrows():
                        sc = fuzz.ratio(t_school, row_s["_norm_school"])
                        if sc > best_score:
                            best_score, best_row_s = sc, row_s

                    # قاعدة التحقق من المدرسة الأصلية
                    school_ok = (
                        best_score >= 85
                        or t_school in best_row_s["_norm_school"]
                        or best_row_s["_norm_school"] in t_school
                    )

                    record["الاسم المطابق"] = best_row_s[name_col_s]
                    record["نسبة تطابق المدرسة"] = f"{round(best_score)}%"
                    record["الملاحظة"] = "✅ اسم + مدرسة" if school_ok else "⚠️ اسم فقط — مدرسة مختلفة"
                    
                    # جلب كافة الأعمدة التي حددها المستخدم ديناميكياً
                    for col in cols_to_fetch:
                        # يتم جلب القيمة فقط إذا تطابقت المدرسة (كما في كودك الأصلي للـ Iban)
                        record[f"جلب_{col}"] = best_row_s[col] if school_ok else ""

                final_results.append(record)

            result_df = pd.DataFrame(final_results)

            # --- التصدير الملون (نفس تنسيق كودك الأصلي) ---
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name="نتائج المطابقة")
                ws = writer.sheets["نتائج المطابقة"]

                green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                # تحديد موقع عمود الملاحظة للتلوين
                note_idx = list(result_df.columns).index("الملاحظة") + 1
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    note_val = str(row[note_idx-1].value)
                    if "✅" in note_val:
                        for cell in row: cell.fill = green_fill
                    elif "⚠️" in note_val:
                        for cell in row: cell.fill = yellow_fill
                    elif "❌" in note_val:
                        for cell in row: cell.fill = red_fill

            st.success("تم الانتهاء! يمكنك الآن تحميل الملف.")
            st.download_button(
                label="📥 تحميل ملف النتائج (Excel)",
                data=output.getvalue(),
                file_name="Matching_Results_Dynamic.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
